import argparse
import json
import os
import openpyxl

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Path to input JSON data")
    parser.add_argument("--template", required=True, help="Path to blank XLSM template")
    parser.add_argument("--output", required=True, help="Path to save output XLSM")
    args = parser.parse_args()

    # Verify input files exist
    if not os.path.exists(args.input):
        print(f"Error: Input JSON {args.input} does not exist.")
        return
    if not os.path.exists(args.template):
        print(f"Error: Template {args.template} does not exist.")
        return

    with open(args.input, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Load template keeping macros
    wb = openpyxl.load_workbook(args.template, keep_vba=True)
    
    if "Template" not in wb.sheetnames:
        print("Error: Template sheet not found in the template workbook.")
        return
        
    sheet = wb["Template"]

    products = data.get("products", [])
    enriched_data = data.get("enrichedData", {})
    global_settings = data.get("globalSettings", {})

    brand = global_settings.get("amazonBrandName") or "CHILS & CO."
    manufacturer = global_settings.get("amazonManufacturer") or "CHILS & CO."
    origin = global_settings.get("amazonOriginCountry") or "India"
    fulfillment = global_settings.get("amazonFulfillmentChannel") or "Merchant Fulfilled"
    handling = global_settings.get("amazonHandlingTime") or "5"
    pkg_length = global_settings.get("amazonDefaultPkgLength") or "35"
    pkg_width = global_settings.get("amazonDefaultPkgWidth") or "25"
    pkg_height = global_settings.get("amazonDefaultPkgHeight") or "5"
    pkg_weight = global_settings.get("amazonDefaultPkgWeight") or "500"
    importer = global_settings.get("amazonImporterContact") or ""
    packer = global_settings.get("amazonPackerContact") or ""

    start_row = 7

    for product in products:
        prod_id = str(product.get("id"))
        enrich = enriched_data.get(prod_id)
        if not enrich:
            continue

        parent_sku = f"P-{prod_id}"

        # Write Parent Row
        parent_cols = {
            1: parent_sku,                                          # SKU
            2: "SHIRT",                                             # Product Type
            3: "full_update",                                       # Listing Action
            4: "parent",                                            # Parentage Level
            6: "SIZE/COLOR",                                        # Variation Theme Name
            7: product.get("name"),                                 # Item Name
            9: brand,                                               # Brand Name
            20: enrich.get("modelName"),                            # Model Name
            21: manufacturer,                                       # Manufacturer
            34: product.get("description"),                         # Product Description
            35: enrich.get("bulletPoints", [""])[0] if len(enrich.get("bulletPoints", [])) > 0 else "",
            36: enrich.get("bulletPoints", [""])[1] if len(enrich.get("bulletPoints", [])) > 1 else "",
            37: enrich.get("bulletPoints", [""])[2] if len(enrich.get("bulletPoints", [])) > 2 else "",
            38: enrich.get("bulletPoints", [""])[3] if len(enrich.get("bulletPoints", [])) > 3 else "",
            39: enrich.get("bulletPoints", [""])[4] if len(enrich.get("bulletPoints", [])) > 4 else "",
            40: enrich.get("genericKeywords"),                      # Generic Keywords
            48: "unisex-adult",                                     # Department Name
            49: "unisex",                                           # Target Gender
            50: "adult",                                            # Age Range Description
            56: product.get("material") or "100% Cotton",           # Material
            59: "French Terry" if enrich.get("apparelFabricWeightClass") == "Heavyweight" else "Cotton Knit", # Fabric Type
            68: parent_sku,                                         # Part Number
            72: product.get("care") or "Machine Wash",              # Care Instructions
            107: importer,                                          # Importer Contact
            112: packer,                                            # Packer Contact
            121: enrich.get("sleeveLengthDescription") or "Long Sleeve", # Sleeve Length Description
            123: enrich.get("closureType") or "Pull-On",             # Closure Type
            133: enrich.get("apparelFabricWeightClass") or "Heavyweight", # Weight Class
            134: "India",                                           # Garment Size Country
            135: 70,                                                # Shoulder to bottom hem length
            136: "Centimeters",                                     # Shoulder to bottom hem length unit
            137: enrich.get("apparelFabricStretch") or "Low Stretch",# Stretch
            138: enrich.get("fitToSizeSentiment") or "True to Size", # Fit Sentiment
            142: enrich.get("itemWeightGrams") or 450,              # Item Weight
            143: "Grams",                                           # Item Weight Unit
            145: "New",                                             # Item Condition
            193: int(pkg_length),                                   # Package Length
            194: "Centimeters",                                     # Package Length Unit
            195: int(pkg_width),                                    # Package Width
            196: "Centimeters",                                     # Package Width Unit
            197: int(pkg_height),                                   # Package Height
            198: "Centimeters",                                     # Package Height Unit
            199: int(pkg_weight),                                   # Package Weight
            200: "Grams",                                           # Package Weight Unit
            201: origin                                             # Country of Origin
        }

        for col_idx, val in parent_cols.items():
            sheet.cell(row=start_row, column=col_idx, value=val)
        start_row += 1

        # Write Child Rows for variations
        variations = product.get("variations", [])
        if variations:
            for v in variations:
                attribs = v.get("attributes", {})
                color = attribs.get("color") or "Oversized"
                size = attribs.get("size") or "Regular"
                child_sku = f"C-{prod_id}-{color.replace(' ', '')}-{size}"
                mapped_color = enrich.get("colorMap", {}).get(color) or "Multicolour"

                child_cols = {
                    1: child_sku,                                           # SKU
                    2: "SHIRT",                                             # Product Type
                    3: "full_update",                                       # Listing Action
                    4: "child",                                             # Parentage Level
                    5: parent_sku,                                          # Parent SKU
                    6: "SIZE/COLOR",                                        # Theme Name
                    7: f"{product.get('name')} (Color: {color}, Size: {size})", # Item Name
                    9: brand,                                               # Brand Name
                    20: enrich.get("modelName"),                            # Model Name
                    21: manufacturer,                                       # Manufacturer
                    34: product.get("description"),                         # Product Description
                    35: enrich.get("bulletPoints", [""])[0] if len(enrich.get("bulletPoints", [])) > 0 else "",
                    36: enrich.get("bulletPoints", [""])[1] if len(enrich.get("bulletPoints", [])) > 1 else "",
                    37: enrich.get("bulletPoints", [""])[2] if len(enrich.get("bulletPoints", [])) > 2 else "",
                    38: enrich.get("bulletPoints", [""])[3] if len(enrich.get("bulletPoints", [])) > 3 else "",
                    39: enrich.get("bulletPoints", [""])[4] if len(enrich.get("bulletPoints", [])) > 4 else "",
                    40: enrich.get("genericKeywords"),                      # Generic Keywords
                    48: "unisex-adult",                                     # Department Name
                    49: "unisex",                                           # Target Gender
                    50: "adult",                                            # Age Range Description
                    51: "ASIAN",                                            # Shirt Size System
                    52: "ALPHA",                                            # Shirt Size Class
                    53: size,                                               # Shirt Size Value
                    55: "Regular",                                          # Shirt Body Type
                    56: product.get("material") or "100% Cotton",           # Material
                    59: "French Terry" if enrich.get("apparelFabricWeightClass") == "Heavyweight" else "Cotton Knit", # Fabric Type
                    65: mapped_color,                                       # Color Map
                    66: color,                                              # Color
                    68: child_sku,                                          # Part Number
                    72: product.get("care") or "Machine Wash",              # Care Instructions
                    107: importer,                                          # Importer Contact
                    112: packer,                                            # Packer Contact
                    121: enrich.get("sleeveLengthDescription") or "Long Sleeve", # Sleeve Length Description
                    123: enrich.get("closureType") or "Pull-On",             # Closure Type
                    133: enrich.get("apparelFabricWeightClass") or "Heavyweight", # Weight Class
                    134: "India",                                           # Garment Size Country
                    135: 70,                                                # Shoulder to bottom hem length
                    136: "Centimeters",                                     # Shoulder to bottom hem length unit
                    137: enrich.get("apparelFabricStretch") or "Low Stretch",# Stretch
                    138: enrich.get("fitToSizeSentiment") or "True to Size", # Fit Sentiment
                    142: enrich.get("itemWeightGrams") or 450,              # Item Weight
                    143: "Grams",                                           # Item Weight Unit
                    145: "New",                                             # Item Condition
                    172: v.get("stockQuantity") or 0,                       # Quantity (IN)
                    173: int(handling),                                     # Handling Time (IN)
                    176: float(v.get("price") or product.get("price")),     # Your Price INR
                    193: int(pkg_length),                                   # Package Length
                    194: "Centimeters",                                     # Package Length Unit
                    195: int(pkg_width),                                    # Package Width
                    196: "Centimeters",                                     # Package Width Unit
                    197: int(pkg_height),                                   # Package Height
                    198: "Centimeters",                                     # Package Height Unit
                    199: int(pkg_weight),                                   # Package Weight
                    200: "Grams",                                           # Package Weight Unit
                    201: origin                                             # Country of Origin
                }

                for col_idx, val in child_cols.items():
                    sheet.cell(row=start_row, column=col_idx, value=val)
                start_row += 1

    wb.save(args.output)
    print("Done")

if __name__ == "__main__":
    main()
